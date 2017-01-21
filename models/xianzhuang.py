from openpyxl import load_workbook


def mianji(value):
    """换算成以“万㎡”为单位"""
    return format(value/1e4, '.2f')


def ratio(value):
    """变换成“涨（跌）xx%”的形式"""
    if value == "NA":
        return "NA"
    elif value < 0:
        return "下降{}%".format(abs(value))
    else:
        return "增长{}%".format(value)


def jine(value):
    """换算成以“亿元”为单位"""
    return format(value / 1e8, '.2f')


def xianzhuang_data():
    """市场运行现状数据
    结构为 { '当月': { '商品房': detail, ...}}
    detail 的结构为{'上市面积': [值, 环比, 同比],'上市套数': [值, 环比, 同比],...}
    """
    def open_sheet(file, sheet):
        """打开表格，返回ws和宽度"""
        wb = load_workbook(file)
        ws = wb[sheet]
        width = ws.max_column
        return ws, width

    def tong_huan_bi(key, start_col):
        for i in [1, 2]:
            data[period][sheet][key].append(ratio(ws.cell(row=row, column=start_col + i).value))

    def get_data():
        """获取数据"""
        # 累计为倒数第17列至第0列，当月为倒数第35列至第18列，都是6*3=18列数据。
        start_col = [width - 17, width - 35][period == '当月']
        # 上市面积
        data[period][sheet]['上市面积'] = list()
        data[period][sheet]['上市面积'].append(mianji(ws.cell(row=row, column=start_col).value))
        tong_huan_bi('上市面积', start_col)
        # 上市套数
        start_col += 3
        data[period][sheet]['上市套数'] = list()
        data[period][sheet]['上市套数'].append(ws.cell(row=row, column=start_col).value)
        tong_huan_bi('上市套数', start_col)
        # 已售面积
        start_col += 3
        data[period][sheet]['已售面积'] = list()
        data[period][sheet]['已售面积'].append(mianji(ws.cell(row=row, column=start_col).value))
        tong_huan_bi('已售面积', start_col)
        # 已售套数
        start_col += 3
        data[period][sheet]['已售套数'] = list()
        data[period][sheet]['已售套数'].append(ws.cell(row=row, column=start_col).value)
        tong_huan_bi('已售套数', start_col)
        # 已售均价
        start_col += 3
        data[period][sheet]['已售均价'] = list()
        data[period][sheet]['已售均价'].append(ws.cell(row=row, column=start_col).value)
        tong_huan_bi('已售均价', start_col)
        # 已售金额
        start_col += 3
        data[period][sheet]['已售金额'] = list()
        data[period][sheet]['已售金额'].append(jine(ws.cell(row=row, column=start_col).value))
        tong_huan_bi('已售金额', start_col)

    file = "E:/lishui/static/lishui1.xlsx"
    row = 4  # 合计项所在行号
    data = dict()
    for period in ['当月', '累计']:
        data[period] = dict()
        for sheet in ['商品房', '住宅', '商办', '商业', '办公']:
            data[period][sheet] = dict()
            ws, width = open_sheet(file, sheet)
            get_data()
    return data

