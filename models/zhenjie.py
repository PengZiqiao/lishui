from openpyxl import load_workbook
import pandas as pd


def mianji(value):
    """换算成以“万㎡”为单位"""
    return float(format(value / 1e4, '.2f'))


def ratio(value):
    """NA都转换成0"""
    return str([value, 0][value == 'NA'])


def to_list(df):
    """把DateFrame转换成按行分格的列表"""
    thead = ['片区', ]
    for each in df.columns:
        thead.append(each)
    tbody = []
    for ix, row in df.iterrows():
        one_row = [ix]
        for each in thead[1:]:
            if row[each] in [0, '-100', '0', '0.00']:
                row[each] = '/'
            elif row[each] == '100.00':
                row[each] = '100'
            one_row.append(row[each])
        tbody.append(one_row)

    return thead, tbody

def zhenjie_data():
    """镇街对比
    生成“二、镇街对比”需要的6张表
    """

    def open_sheet(sheet):
        """打开表格，返回worksheet、宽度和一个DateFrame对象"""
        wb = load_workbook(file)
        worksheet = wb[sheet]
        max_col = worksheet.max_column
        return worksheet, max_col

    def gong_ying():
        df = pd.DataFrame(index=index)
        i = start_row
        for ix, row in df.iterrows():
            for period in ['当月', '累计']:
                # 累计为倒数第6列至第0列，当月为倒数第13列至第7列，分别有7列
                col = [width - 6, width - 13][period == '当月']
                df.at[ix, period + '面积'] = mianji(ws.cell(row=i, column=col).value)
                if period == '当月':
                    df.at[ix, period + '环比'] = ratio(ws.cell(row=i, column=col + 1).value)
                df.at[ix, period + '同比'] = ratio(ws.cell(row=i, column=col + 2).value)
                if df.at['合计', period + '面积'] == 0:
                    df.at[ix, period + '占比'] = 0
                else:
                    df.at[ix, period + '占比'] = str(
                        format(ws.cell(row=i, column=col).value / ws.cell(row=start_row, column=col).value * 100, '.2f'))
            i += 1
        return to_list(df.reindex(new_index))

    def cheng_jiao():
        df = pd.DataFrame(index=index)
        i = start_row
        for ix, row in df.iterrows():
            for period in ['当月', '累计']:
                # 累计为倒数第6列至第0列，当月为倒数第13列至第7列，已售从每块的第4列开始
                col = [width - 3, width - 10][period == '当月']
                df.at[ix, period + '面积'] = mianji(ws.cell(row=i, column=col).value)
                if period == '当月':
                    df.at[ix, period + '环比'] = ratio(ws.cell(row=i, column=col + 1).value)
                df.at[ix, period + '同比'] = ratio(ws.cell(row=i, column=col + 2).value)
                if df.at['合计', period + '面积'] == 0:
                    df.at[ix, period + '占比'] = 0
                else:
                    df.at[ix, period + '占比'] = str(
                        format(ws.cell(row=i, column=col).value / ws.cell(row=start_row, column=col).value * 100, '.2f'))
                if period == '当月':
                    df.at[ix, period + '均价'] = str(ws.cell(row=i, column=col + 3).value)
            i += 1
        return to_list(df.reindex(new_index))

    file = "E:/lishui/static/lishui2.xlsx"
    index = ['合计', '永阳街道', '开发区', '洪蓝镇', '白马镇', '和凤镇', '东屏镇', '石湫镇', '晶桥镇']
    new_index = ['永阳街道', '开发区', '洪蓝镇', '白马镇', '和凤镇', '东屏镇', '石湫镇', '晶桥镇', '合计']
    sheet_list = ['住宅', '商业', '办公']
    start_row = 4
    data = dict()
    for each in sheet_list:
        data[each] = list()
        ws, width = open_sheet(each)
        data[each].append(gong_ying())
        data[each].append(cheng_jiao())
    return data
